import json
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpRequest, HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login, logout
from django.contrib import messages
# # create_accounts_from_excel.py
from PDF_processing.models import Comment, Respon, Section, Teacher, TeacherCourse,Course
import openpyxl
from django.contrib.auth.models import User
import random
import string
import PyPDF2
from requests import Response
from transformers import pipeline
from docx import Document
import warnings
from django.contrib.auth.models import Group
from django.http import JsonResponse
global_course=""
# Suppress the specific warning
warnings.filterwarnings("ignore", category=PendingDeprecationWarning)
#from django.contrib.auth.decorators import login_required
#from django.contrib import messages
#from django.shortcuts import render, redirect
from django.contrib.auth import update_session_auth_hash
# Create your views here.
from django import forms

from django import forms
from django import forms

import tkinter as tk
from tkinter import filedialog
import shutil
import os



def home(request):
   
    return render(request, 'authentication\index.html')
    
def signup(request):
    if request.method=="POST":
        username=request.POST['username']
        fname=request.POST['firstname']
        lname=request.POST['lastname']
        email=request.POST['email']
        pass1=request.POST['password']
        pass2=request.POST['confirm-password']
        selected_role = request.POST['role']
        if User.objects.filter(username=username):
            messages.error(request, "Username already exist! Please try some other username.")
            return redirect('home')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email Already Registered!!")
            return redirect('home')
        
        if len(username)>20:
            messages.error(request, "Username must be under 20 charcters!!")
            return redirect('home')
        
        if pass1 != pass2:
            messages.error(request, "Passwords didn't matched!!")
            return redirect('home')
        
        if not username.isalnum():
            messages.error(request, "Username must be Alpha-Numeric!!")
            return redirect('home')
        myuser=User.objects.create_user(username, email,pass1)
        myuser.first_name=fname
        myuser.last_name=lname
        myuser.save()
        user = User.objects.get(username=username)
        group, created = Group.objects.get_or_create(name='HOS')
        group, created = Group.objects.get_or_create(name='HOD CS')
        group, created = Group.objects.get_or_create(name='HOD SE')
        group, created = Group.objects.get_or_create(name='HOD EE')
        group, created = Group.objects.get_or_create(name='HOD Humanit')
        group, created = Group.objects.get_or_create(name='HOD FSM')
        group, created = Group.objects.get_or_create(name='Admin')
        #user.groups.add(instructor_group)
        user.groups.add(group)
        messages.success(request,"the user has been successfuly registered")
        return redirect('signin')
    return render(request, 'authentication\signup.html')
    
    
        # <input type="email" id="email" name="email" required>
        
        # <label for="firstname">First Name:</label>
        # <input type="text" id="firstname" name="firstname" required>
        
        # <label for="lastname">Last Name:</label>
        # <input type="text" id="lastname" name="lastname" required>
        
        # <label for="username">Username:</label>
        # <input type="text" id="username" name="username" required>
        
        # <label for="password">Password:</label>
        # <input type="password" id="password" name="password" required>
        
        # <label for="confirm-password">Confirm Password:</label>
        # <input type="password" id="confirm-password" name="confirm-password" required>
        
def is_admin(user):
    return user.groups.filter(name='admin').exists()

from rest_framework.views import APIView
from rest_framework import status
# class SignInView (APIView):
#     def post(self, request, format=None):
#         username = request.data.get('username')
#         password = request.data.get('password')
#         user=authenticate(username=username, password=password)

#         if username == valid_username and password == valid_password:
#             return Response({'message': 'Login successful'}, status=status.HTTP_200_OK)
#         else:
#             return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)



# from django.shortcuts import render

def home(request):
    return render(request, 'authentication\\index.html')  # Assuming you have a 'home.html' template

# def signin(request):
#     if request.method=="POST":
#         username=request.POST['username']
#         pass1=request.POST['password']
#         user=authenticate(username=username, password=pass1)
#         if user is not None:
#             login(request, user)
#             messages.success(request,"you have logged in")
#             if(is_admin(user)):
#                 return render(request,'authentication\\index.html',{'fname':user.first_name} )
#             else:
#                 return render(request,'authentication\\not_admin.html',{'fname':user.first_name} )
#         else:
#             messages.error(request,"bad credentials")
#             return redirect('home')



#     return render(request, 'authentication\\signin.html')
# class SignInView(APIView):
#     def post(self, request):
#         username = request.data.get('username')
#         pass1 = request.data.get('password')
#         user = authenticate(username=username, password=pass1)
#         if user is not None:
            # login(request, user)
            # if is_admin(user):
            #     return Response({'message': 'You have logged in', 'fname': user.first_name}, status=status.HTTP_200_OK)
            # else:
            #     return Response({'message': 'You have logged in as a non-admin user', 'fname': user.first_name}, status=status.HTTP_200_OK)
#         else:
#             messages.error(request, "Bad credentials")
#             return Response({'error': 'Bad credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework.response import Response

def is_user_in_group(user, group_name):
    try:
        group = Group.objects.get(name=group_name)
        return True
    except Group.DoesNotExist:
        return False

@api_view(['POST','GET'])
def setcourse(request):
    try:
        course = request.data.get('selectedCourse')
        # Assuming you want to store the selected course globally
        global global_course
        global_course = course

        # Your additional logic with the selected course, if needed

        # Create a JSON response
        response_data = {
            'status': 'success',
            'message': 'Course set successfully',
            'selectedCourse': global_course,  # Include the selected course in the response if needed
        }
    
        return JsonResponse(response_data, safe=False)
    except Exception as e:
        # Handle exceptions appropriately
        response_data = {
            'status': 'error',
            'message': f'Error: {str(e)}',
        }
        return JsonResponse(response_data, status=500, safe=False)
# @api_view(['POST','GET'])
# def feedbacksReceived(request):
#     try:
#         print("\n\nglobal to")
#         print(global_to,"\n\n\n\n")
#         if global_to=='' and global_from=='':
#             print("we arre in condition 1")
#             teacher = Teacher.objects.get(first_name=set_g_instructor)
#             print(teacher)
#             print("reach cond 2")
#             # Get the Course object
#             # course = Course.objects.get(course_name=global_course)
#             # print(course)
#             # Get the TeacherCourse object
#             teacher_course = TeacherCourse.objects.filter(teacher=teacher)
#             for tech in teacher_course:
#             # Get all sections related to the TeacherCourse
#                 sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
#                 feedbacksrecieved = 0
#                 for sec in sections:
#                     print(sec.section_name)
#                     print(sec.semester)
#                     print(sec.year)
#                     print("section feed backes")
#                     print(sec.feedbacks_received)
#                     # feedbacksrecieved = sec.feedbacks_received
#                     feedbacksrecieved = feedbacksrecieved+int(sec.feedbacks_received)
#                     print(feedbacksrecieved)
#                     print("added")
                    

#             # Initialize a variable to store the total registered students
#             # feedbacksrecieved = 0
#             # print(sections)
#             # # Loop through each section and add the registered students count
#             # for section in sections:
#             #     feedbacksrecieved += section.feedbacks_received

#             # Create a JSON response
#             response_data = {'value': feedbacksrecieved}

#             return JsonResponse(response_data)

#     except Teacher.DoesNotExist:
#         return JsonResponse({'error': 'Teacher not found'})

#     except Course.DoesNotExist:
#         return JsonResponse({'error': 'Course not found'})

#     except TeacherCourse.DoesNotExist:
#         return JsonResponse({'error': 'TeacherCourse not found'})
global_section=""
@api_view(['POST','GET'])
def feedbacksReceived(request):
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                for sec in sections:
                    feedbacks_received += int(sec.feedbacks_received)
            
            response_data = {'value': feedbacks_received}
            print("\n\n\nresponse_data: ",response_data)
            print("\n\n\n\n")
            return JsonResponse(response_data)
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                for sec in sections:
                    feedbacks_received += int(sec.feedbacks_received)

            response_data = {'value': feedbacks_received}
            return JsonResponse(response_data)

            
            
        else:
            # print("we in third condition")
            From= global_from
            To= global_to
            set_g_instructor 
            # = request.data.get('selectedInstructor')
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            # set_g_instructor = request.data.get('selectedInstructor')
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)
            start_semester, start_year = From.split()
            end_semester, end_year = To.split()
            end_year=int(end_year)
            start_year=int(start_year)
            SEMESTER_ORDER = {
            'spring': 1,
            'summer': 2,
            'fall': 3,
            }

            years_range = list(range(start_year, end_year + 1))


            sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
            )
            # Get unique courses associated with filtered sections
            # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
            # print(sections)
            
            total_registered_students = 0

                # Loop through each section and add the registered students count
            for section in sections:
                # print(section)
                # print(section.number_of_positive_comments)
                # print("\n\n\n")
                total_registered_students += section.feedbacks_received

            # Create a JSON response
            response_data = {'value': total_registered_students}

            return JsonResponse(response_data)
            # Handle the case when global_to or global_from are not empty
            # You may want to return an appropriate response or handle the case differently
            pass

    except Teacher.DoesNotExist:
        # Handle the case when the teacher doesn't exist
        pass

    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})

@api_view(['POST','GET'])
def getcomments(request):
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            all_comments=[]
            for tech in teacher_course:
                if(global_section==""):
                    print("\n\n\nthe sections")
                    
                    sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                    print(sections)
                    print("something")
                    comments = Comment.objects.filter(section__in=sections)
                    print("\n\n\n\n")
                    print(comments[0].comments)
                    print("these are the comments")
                    print(global_cat)
                    if global_cat!=None :
                        print("global_cat  is not none")
                        comments = comments.filter(category=global_cat)

                    print(comments)
                    for com in comments: 
                        print(com.comments)
                        all_comments.append(com.comments)
                    data={
                    'comments_data':all_comments
                }
                    return JsonResponse(data, safe=False)

                else:
                    sections = Section.objects.filter(section_name=global_section,teacher_course=tech, semester=g_semester, year=g_year)
                    comments = Comment.objects.filter(section__in=sections)
                    if global_cat!=None:
                        comments = comments.filter(category=global_cat)

                    
                    for com in comments: 
                        all_comments.append(com.comments)
                
                # Convert queryset to a list of dictionaries
                # comments_data = serialize('json', comments)
                data={
                    'comments_data':all_comments
                }
                return JsonResponse(data, safe=False)
            
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            all_comments=[]
            for tech in teacher_course:
                if(global_section==""):
                    print("\n\n\nthe sections")
                    
                    sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                    print(sections)
                    print("something")
                    comments = Comment.objects.filter(section__in=sections)
                    print("\n\n\n\n")
                    print(comments[0].comments)
                    print("these are the comments")
                    print(global_cat)
                    if global_cat!=None :
                        print("global_cat  is not none")
                        comments = comments.filter(category=global_cat)

                    print(comments)
                    for com in comments: 
                        print(com.comments)
                        all_comments.append(com.comments)
                    data={
                    'comments_data':all_comments
                }
                    return JsonResponse(data, safe=False)

                else:
                    sections = Section.objects.filter(section_name=global_section,teacher_course=tech, semester=g_semester, year=g_year)
                    comments = Comment.objects.filter(section__in=sections)
                    if global_cat!=None:
                        comments = comments.filter(category=global_cat)

                    
                    for com in comments: 
                        all_comments.append(com.comments)
                
                # Convert queryset to a list of dictionaries
                # comments_data = serialize('json', comments)
                data={
                    'comments_data':all_comments
                }
                return JsonResponse(data, safe=False)

            
            
        else:
            # print("we in third condition")
            From= global_from
            To= global_to
            set_g_instructor 
            # = request.data.get('selectedInstructor')
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            # set_g_instructor = request.data.get('selectedInstructor')
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)
            start_semester, start_year = From.split()
            end_semester, end_year = To.split()
            end_year=int(end_year)
            start_year=int(start_year)
            SEMESTER_ORDER = {
            'spring': 1,
            'summer': 2,
            'fall': 3,
            }

            years_range = list(range(start_year, end_year + 1))


            sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
            )
            # Get unique courses associated with filtered sections
            # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
            # print(sections)
            if global_section!="":
                sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course, section_name=global_section
            )
            
            print(sections)    
           
            comments = Comment.objects.filter(section__in=sections)
            if  global_cat != None:
                comments = comments.filter(category=global_cat)

            all_comments=[]
            for com in comments: 
                all_comments.append(com.comments)

            # Convert queryset to a list of dictionaries
            # comments_data = serialize('json', comments)
            data={
                'comments_data':all_comments
            }
            return JsonResponse(data, safe=False)

            # return JsonResponse(response_data)
            # Handle the case when global_to or global_from are not empty
            # You may want to return an appropriate response or handle the case differently
            pass

    except Teacher.DoesNotExist:
        # Handle the case when the teacher doesn't exist
        pass

    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})

    



global_course=""
@api_view(['POST'])
def setSection(request):
    try:
        section = request.data.get('selectedSection')
        # Assuming you want to store the selected course globally
        global global_section
        global_section = section

        # Your additional logic with the selected course, if needed

        # Create a JSON response
        response_data = {
            'status': 'success',
            'message': 'Course set successfully',
            'selectedCourse': global_section,  # Include the selected course in the response if needed
        }
    
     
        return JsonResponse(response_data, safe=False)
    except Exception as e:
        # Handle exceptions appropriately
        response_data = {
            'status': 'error',
            'message': f'Error: {str(e)}',
        }
        return JsonResponse(response_data, status=500, safe=False)
global_cat=""
@api_view(['POST','GET'])
def setCategory(request):
    try:
        cat = request.data.get('selectedCategory')
        # global_cat=request.data.get('selectedCategory')

        # Assuming you want to store the selected course globally
        global global_cat
        global_cat = cat
        print("\n\n\nselectedcat: ",global_cat)
        print("\n\n\n")
        response_data = {
            'status': 'success',
            'message': 'Category set successfully',
            'selectedCourse': global_cat,  
        }
        # print("global_cat\n\n\n",global_cat)
        return JsonResponse(response_data, safe=False)
    except Exception as e:
        # Handle exceptions appropriately
        response_data = {
            'status': 'error',
            'message': f'Error: {str(e)}',
        }
        return JsonResponse(response_data, status=500, safe=False)


@api_view(['POST','GET'])
def course_analysis(request):
    try:
        set_g_instructor
        # Get the Teacher object
        teacher = Teacher.objects.get(first_name=set_g_instructor)

        # Get the TeacherCourse object
        teacher_course = TeacherCourse.objects.filter(teacher=teacher)
        courses_name = []
        positive_ones = []
        negative_ones = []

        for teach in teacher_course:
            courses_name.append(teach.course.course_name)
            positive = 0
            negative = 0

            all_sections_of_subject = Section.objects.filter(teacher_course=teach)
            
            for section in all_sections_of_subject:
                positive += section.number_of_positive_comments
                negative += section.number_of_constructive_comments

            positive_ones.append(positive)
            negative_ones.append(negative)

        data = {
            'positive': positive_ones,
            'negative': negative_ones,
            'course_name': courses_name,
        }

        return JsonResponse(data)

    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'})

    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'})

    except TeacherCourse.DoesNotExist:
        return JsonResponse({'error': 'TeacherCourse not found'})

        # return JsonResponse({'error': 'TeacherCourse not found'})
@api_view(['POST','GET'])
def negcomments(request):
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                for sec in sections:
                    feedbacks_received += int(sec.number_of_constructive_comments)
            
            response_data = {'value': feedbacks_received}
            print("\n\n\nresponse_data: ",response_data)
            print("\n\n\n\n")
            return JsonResponse(response_data)
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                for sec in sections:
                    feedbacks_received += int(sec.number_of_constructive_comments)

            response_data = {'value': feedbacks_received}
            return JsonResponse(response_data)

            
            
        else:
            # print("we in third condition")
            From= global_from
            To= global_to
            set_g_instructor 
            # = request.data.get('selectedInstructor')
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            # set_g_instructor = request.data.get('selectedInstructor')
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)
            start_semester, start_year = From.split()
            end_semester, end_year = To.split()
            end_year=int(end_year)
            start_year=int(start_year)
            SEMESTER_ORDER = {
            'spring': 1,
            'summer': 2,
            'fall': 3,
            }

            years_range = list(range(start_year, end_year + 1))


            sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
            )
            # Get unique courses associated with filtered sections
            # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
            # print(sections)
            
            total_registered_students = 0

                # Loop through each section and add the registered students count
            for section in sections:
                # print(section)
                # print(section.number_of_positive_comments)
                # print("\n\n\n")
                total_registered_students += section.number_of_constructive_comments

            # Create a JSON response
            response_data = {'value': total_registered_students}

            return JsonResponse(response_data)
            # Handle the case when global_to or global_from are not empty
            # You may want to return an appropriate response or handle the case differently
            pass

    except Teacher.DoesNotExist:
        # Handle the case when the teacher doesn't exist
        pass

    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})
    # try:
        
    #     teacher = Teacher.objects.get(first_name=set_g_instructor)

    #     # Get the Course object
    #     course = Course.objects.get(course_name=global_course)

    #     # Get the TeacherCourse object
    #     teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)

    #     # Get all sections related to the TeacherCourse
    #     sections = Section.objects.filter(teacher_course=teacher_course)

    #     # Initialize a variable to store the total registered students
    #     criticized = 0

    #     # Loop through each section and add the registered students count
    #     for section in sections:
    #         criticized += section.number_of_constructive_comments

    #     # Create a JSON response
    #     response_data = {'value': criticized}

    #     return JsonResponse(response_data)
    # except Teacher.DoesNotExist:
    #     return JsonResponse({'error': 'Teacher not found'})

    # except Course.DoesNotExist:
    #     return JsonResponse({'error': 'Course not found'})

    # except TeacherCourse.DoesNotExist:
    #     return JsonResponse({'error': 'TeacherCourse not found'})
    



# @api_view(['POST','GET'])
# def category_wise(request):
#     try:
       
#         teacher = Teacher.objects.get(first_name=set_g_instructor)

#         # Get the Course object
#         course = Course.objects.get(course_name=global_course)

#         # Get the TeacherCourse object
#         teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)

#         # Get all sections related to the TeacherCourse
#         sections = Section.objects.filter(teacher_course=teacher_course)

#         # Initialize a variable to store the total registered students
#         criticized = 0

#         # Loop through each section and add the registered students count
#         for section in sections:
#             criticized += section.number_of_constructive_comments

#         # Create a JSON response
#         response_data = {'value': criticized}

#         return JsonResponse(response_data)
#     except Teacher.DoesNotExist:
#         return JsonResponse({'error': 'Teacher not found'})

#     except Course.DoesNotExist:
#         return JsonResponse({'error': 'Course not found'})

#     except TeacherCourse.DoesNotExist:
#         return JsonResponse({'error': 'TeacherCourse not found'})
@api_view(['POST','GET'])
def Constructivecomments(request):
   
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                for sec in sections:
                    feedbacks_received += int(sec.number_of_positive_comments)
            
            response_data = {'value': feedbacks_received}
            print("\n\n\nresponse_data: ",response_data)
            print("\n\n\n\n")
            return JsonResponse(response_data)
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                for sec in sections:
                    feedbacks_received += int(sec.number_of_positive_comments)

            response_data = {'value': feedbacks_received}
            return JsonResponse(response_data)

            
            
        else:
            # print("we in third condition")
            From= global_from
            To= global_to
            set_g_instructor 
            # = request.data.get('selectedInstructor')
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            # set_g_instructor = request.data.get('selectedInstructor')
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)
            start_semester, start_year = From.split()
            end_semester, end_year = To.split()
            end_year=int(end_year)
            start_year=int(start_year)
            SEMESTER_ORDER = {
            'spring': 1,
            'summer': 2,
            'fall': 3,
            }

            years_range = list(range(start_year, end_year + 1))


            sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
            )
            # Get unique courses associated with filtered sections
            # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
            # print(sections)
            
            total_registered_students = 0

                # Loop through each section and add the registered students count
            for section in sections:
                # print(section)
                # print(section.number_of_positive_comments)
                # print("\n\n\n")
                total_registered_students += section.number_of_positive_comments

            # Create a JSON response
            response_data = {'value': total_registered_students}

            return JsonResponse(response_data)
            # Handle the case when global_to or global_from are not empty
            # You may want to return an appropriate response or handle the case differently
            pass

    except Teacher.DoesNotExist:
        # Handle the case when the teacher doesn't exist
        pass

    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})

@api_view(['POST','GET'])
def commentscategorywise(request):
    try:
        # Get the Teacher object
      
        teacher = Teacher.objects.get(first_name=set_g_instructor)

        # Get the Course object
        course = Course.objects.get(course_name=global_course)

        # Get the TeacherCourse object
        teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)

        # Get all sections related to the TeacherCourse
        sections = Section.objects.filter(teacher_course=teacher_course)

        # Initialize a variable to store the total registered students
        positive = 0

        # Loop through each section and add the registered students count
        learning_material_count = 0
        course_delivery_count = 0
        something1_count = 0
        something2_count = 0
        sum_positive_comments_ass = 0
        sum_positive_comments_content_org = 0
        sum_positive_comments_learning_mat = 0
        sum_positive_comments_classenv = 0
        for section in sections:
            sum_positive_comments_ass += section.number_of_positive_comments_ass
            sum_positive_comments_content_org += section.number_of_positive_comments_content_org
            sum_positive_comments_learning_mat += section.number_of_positive_comments_learning_mat
            sum_positive_comments_classenv += section.number_of_positive_comments_classenv

        for section in sections:
            

    # Get comments for the specified Section ID
            comments = Comment.objects.filter(section_id=section.id)

            # Iterate through comments and count for each category
            for comment in comments:
                if comment.category == 'Instructor’s quality of delivery of lectures and classroom learning environment':
                    learning_material_count += 1
                elif comment.category == 'Course Content and Organization':
                    course_delivery_count += 1
                elif comment.category == 'Assignment, Quizzes  Evaluation':
                    something1_count += 1
                elif comment.category == 'Learning Material (Textbook, References Books, Videos etc)':
                    something2_count += 1
        positive=[sum_positive_comments_ass,sum_positive_comments_content_org,sum_positive_comments_learning_mat,sum_positive_comments_classenv]
    # Return a dictionary with counts for each category
        negative=[learning_material_count,course_delivery_count,something1_count,something2_count]
       
        data={
              
                'positive':positive,
                'negative':negative
                }
          # 'quality_of_lec': learning_material_count,
                # 'course_content': course_delivery_count,
                # 'quiz': something1_count,
                # 'learningmat': something2_count,
                
                # "Sum_of_number_of_positive_comments_ass": sum_positive_comments_ass,
                # "Sum_of_number_of_positive_comments_content_org": sum_positive_comments_content_org,
                # "Sum_of_number_of_positive_comments_learning_mat": sum_positive_comments_learning_mat,
                # "Sum_of_number_of_positive_comments_classenv": sum_positive_comments_classenv,

        # Create a JSON response
        # response_data = {'value': positive}

        return JsonResponse(data)

    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'})

    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'})

    except TeacherCourse.DoesNotExist:
        return JsonResponse({'error': 'TeacherCourse not found'})
    # 
@api_view(['POST','GET'])
def get_registered_students_json(request):
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                for sec in sections:
                    feedbacks_received += int(sec.registered_students)
            
            response_data = {'value': feedbacks_received}
            print("\n\n\nresponse_data: ",response_data)
            print("\n\n\n\n")
            return JsonResponse(response_data)
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                for sec in sections:
                    feedbacks_received += int(sec.registered_students)

            response_data = {'value': feedbacks_received}
            return JsonResponse(response_data)

            
            
        else:
            # print("we in third condition")
            From= global_from
            To= global_to
            set_g_instructor 
            # = request.data.get('selectedInstructor')
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            # set_g_instructor = request.data.get('selectedInstructor')
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)
            start_semester, start_year = From.split()
            end_semester, end_year = To.split()
            end_year=int(end_year)
            start_year=int(start_year)
            SEMESTER_ORDER = {
            'spring': 1,
            'summer': 2,
            'fall': 3,
            }

            years_range = list(range(start_year, end_year + 1))


            sections = Section.objects.filter(
                Q(year__in=years_range) |
                Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
                Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
                Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
            )
            # Get unique courses associated with filtered sections
            # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
            # print(sections)
            
            total_registered_students = 0

                # Loop through each section and add the registered students count
            for section in sections:
                # print(section)
                # print(section.number_of_positive_comments)
                # print("\n\n\n")
                total_registered_students += section.registered_students

            # Create a JSON response
            response_data = {'value': total_registered_students}

            return JsonResponse(response_data)
            # Handle the case when global_to or global_from are not empty
            # You may want to return an appropriate response or handle the case differently
            pass

    except Teacher.DoesNotExist:
        # Handle the case when the teacher doesn't exist
        pass

    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})
    # try:
    #     set_g_instructor 
    #     print("srtginstruct:\n\n\n",set_g_instructor)
    #     teacher = Teacher.objects.get(first_name=set_g_instructor)
    #     # = request.data.get('selectedInstructor')
    #     # Ensure global variables are defined properly
    #     global global_to
    #     global global_from

    #     if ('global_to' not in globals() or 'global_from' not in globals() or
    #         global_to == '' or global_to is None or global_from == '' or global_from is None):
    #         # Assuming set_g_instructor is defined somewhere else
            
    #         teacher_course=""
    #         if global_course!="":
    #             course = Course.objects.get(course_name=global_course)
    #             # Assuming g_semester and g_year are defined somewhere else
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
    #         else:
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher)

    #         feedbacks_received = 0
            
    #         for tech in teacher_course:
    #             sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
    #             for sec in sections:
    #                 feedbacks_received += int(sec.registered_students)
            
    #         response_data = {'value': feedbacks_received}
    #         print("\n\n\nresponse_data: ",response_data)
    #         print("\n\n\n\n")
    #     elif global_to == global_from:
    #         sem,yea=global_to.split()
    #         teacher_course=""
    #         if global_course!="":
    #             course = Course.objects.get(course_name=global_course)
    #             # Assuming g_semester and g_year are defined somewhere else
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
    #         else:
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher)

    #         feedbacks_received = 0
            
    #         for tech in teacher_course:
    #             sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
    #             for sec in sections:
    #                 feedbacks_received += int(sec.registered_students)

    #         response_data = {'value': feedbacks_received}
    #         return JsonResponse(response_data)

            
            
    #     else:
    #         # print("we in third condition")
    #         From= global_from
    #         To= global_to
    #         set_g_instructor = request.data.get('selectedInstructor')
    #         teacher = Teacher.objects.get(first_name=set_g_instructor)
    #         teacher_course=""
    #         # set_g_instructor = request.data.get('selectedInstructor')
    #         if global_course!="":
    #             course = Course.objects.get(course_name=global_course)
    #             # Assuming g_semester and g_year are defined somewhere else
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
    #         else:
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher)
    #         start_semester, start_year = From.split()
    #         end_semester, end_year = To.split()
    #         end_year=int(end_year)
    #         start_year=int(start_year)
    #         SEMESTER_ORDER = {
    #         'spring': 1,
    #         'summer': 2,
    #         'fall': 3,
    #         }

    #         years_range = list(range(start_year, end_year + 1))


    #         sections = Section.objects.filter(
    #             Q(year__in=years_range) |
    #             Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
    #             Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
    #             Q(year__gt=start_year, year__lt=end_year), teacher_course__in=teacher_course
    #         )
    #         # Get unique courses associated with filtered sections
    #         # filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
            
                
           
    #         # print(sections)
            
    #         total_registered_students = 0

    #             # Loop through each section and add the registered students count
    #         for section in sections:
    #             # print(section)
    #             # print(section.number_of_positive_comments)
    #             # print("\n\n\n")
    #             total_registered_students += section.registered_students

    #         # Create a JSON response
    #         response_data = {'value': total_registered_students}

    #         return JsonResponse(response_data)
    #         # Handle the case when global_to or global_from are not empty
    #         # You may want to return an appropriate response or handle the case differently
    #         pass

    # except Teacher.DoesNotExist:
    #     # Handle the case when the teacher doesn't exist
    #     pass

    # except Exception as e:
    #     # Handle other exceptions
    #     print(e)
    #     pass

    # # Return a default response if conditions are not met or if an error occurs
    # return JsonResponse({'error': 'Could not calculate feedbacks received'})
    
    # try:
    #     set_g_instructor = request.data.get('selectedInstructor')
    #     if request.data.get('selectedFrom')==None and request.data.get('selectedTo')==None:
    #         print("we in the firs tcondition")
    #         set_g_instructor = request.data.get('selectedInstructor')
    #         teacher = Teacher.objects.get(first_name=set_g_instructor)
    #         if global_course=="":
    #             teacher = Teacher.objects.get(first_name=set_g_instructor)
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher)
    #             total_registered_students = 0
    #             for tech in teacher_course:
    #                 sections = Section.objects.filter(teacher_course=tech)
    #                 for sect in sections:
    #                     total_registered_students += sect.registered_students
    #             response_data = {'value': total_registered_students}
    #             return JsonResponse(response_data)
                                                
    #         else:
    #             print("first condition with course")
    #             print("global_course\n",global_course)
    #             course = Course.objects.get(course_name=global_course)
    #             teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)
    #             total_registered_students = 0
    #             sections = Section.objects.filter(teacher_course=teacher_course)
    #             for sect in sections:
    #                 total_registered_students += sect.registered_students
    #             response_data = {'value': total_registered_students}
    #             return JsonResponse(response_data)

    #     elif request.data.get('selectedFrom')==request.data.get('selectedTo'):
    #         print("\n\n\nsame range\n\n\n")
    #         if global_course=="":
    #             print("\n\n\nsame range without curse\n\n\n")
    #             From=request.data.get('selectedFrom')
    #             gsem,gyear=From.split()
    #             set_g_instructor = request.data.get('selectedInstructor')
    #             teacher = Teacher.objects.get(first_name=set_g_instructor)
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher)
    #             sections = Section.objects.filter(teacher_course=teacher_course, year=gyear, semester=gsem)
    #             total_registered_students = 0

    #         # Loop through each section and add the registered students count
    #             for section in sections:
    #                 total_registered_students += section.registered_students

    #             # Create a JSON response
    #             response_data = {'value': total_registered_students}

    #             return JsonResponse(response_data)
    #         else:
    #             print("\n\n\nsame range with curse\n\n\n")
    #             From=request.data.get('selectedFrom')
    #             gsem,gyear=From.split()
    #             set_g_instructor = request.data.get('selectedInstructor')
    #             teacher = Teacher.objects.get(first_name=set_g_instructor)
    #             print("teacher gottend")
    #             teacher_course = TeacherCourse.objects.filter(teacher=teacher, course_id=global_course)
    #             print("teacher course gotten")
    #             sections = Section.objects.filter(teacher_course=teacher_course, year=gyear, semester=gsem)
    #             print(sections)
    #             total_registered_students = 0

    #             # Loop through each section and add the registered students count
    #             for section in sections:
    #                 total_registered_students += section.registered_students

    #             # Create a JSON response
    #             response_data = {'value': total_registered_students}

    #             return JsonResponse(response_data)
    #     else:
    #         print("\n\n\nrange \n\n\n")
    #         From= request.data.get('selectedFrom')
    #         To= request.data.get('selectedTo')
    #         set_g_instructor = request.data.get('selectedInstructor')
    #         # set_g_instructor = request.data.get('selectedInstructor')
    #         course = Course.objects.get(course_name=global_course)
    #         start_semester, start_year = From.split()
    #         end_semester, end_year = To.split()
    #         end_year=int(end_year)
    #         start_year=int(start_year)
    #         SEMESTER_ORDER = {
    #         'spring': 1,
    #         'summer': 2,
    #         'fall': 3,
    #         }

    #         years_range = list(range(start_year, end_year + 1))


    #         sections = Section.objects.filter(
    #             Q(year__in=years_range) |
    #             Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
    #             Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
    #             Q(year__gt=start_year, year__lt=end_year)
    #         )
    #         # Get unique courses associated with filtered sections
    #         filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
    #         if global_course!="":
    #             print("\n\n\nrange with curse\n\n\n")
    #             filtered_sections = filtered_sections.filter(teacher_course__course__course_name=global_course)
    #         else:
    #              print("\n\n\nsame range without curse\n\n\n")

    #         sections=filtered_sections
    #         total_registered_students = 0

    #             # Loop through each section and add the registered students count
    #         for section in sections:
    #             total_registered_students += section.registered_students

    #         # Create a JSON response
    #         response_data = {'value': total_registered_students}

    #         return JsonResponse(response_data)





    # except Teacher.DoesNotExist:
    #     return JsonResponse({'error': 'Teacher not found'})

    # except Course.DoesNotExist:
    #     return JsonResponse({'error': 'Course not found'})

    # except TeacherCourse.DoesNotExist:
    #     return JsonResponse({'error': 'TeacherCourse not found'})
   
g_year =""
g_semester=""
@api_view(['POST'])
def signin(request):
    if request.method == "POST":
        username = request.data.get('username')
        password = request.data.get('password')
        global g_year # assign to current
        global g_semester # assign to current 
        g_semester = 'Spring'
        g_year = '2022'
        user = Teacher.objects.get(email=username)
       
        global set_g_instructor
          # Declare that you are using the global variable
        global logged_in
        logged_in=user.first_name
        set_g_instructor=user.first_name
        try:
            # user = Teacher.objects.get(Q(email=username) & Q(password=password))
            user_groups = user.groups.all()
            if any(group.name == 'HOS' for group in user_groups):
                return Response({'role': 'HOS','first_name':set_g_instructor} ,status=status.HTTP_200_OK)
            if any(group.name == 'Admin' for group in user_groups):
                return Response({'role': 'Admin','first_name':set_g_instructor} ,status=status.HTTP_200_OK)
            else:
                return Response({'role': 'teacher','first_name':set_g_instructor} ,status=status.HTTP_200_OK)
            
                
        except :
            return json.dumps({"error": "User not found"})
@api_view(['POST'])
def check_HOS(request):
    global set_g_instructor 
    user = Teacher.objects.get(first_name=set_g_instructor )
#     global set_g_instructor  # Declare that you are using the global variable
#     set_g_instructor=user.first_name
    try:
        # user = Teacher.objects.get(Q(email=username) & Q(password=password))
        user_groups = user.groups.all()
        if any(group.name == 'HOS' for group in user_groups):
            return Response({'role': 'HOS'} ,status=status.HTTP_200_OK)
        if any(group.name == 'Admin' for group in user_groups):
            return Response({'role': 'Admin'} ,status=status.HTTP_200_OK)
        else:
            return Response({'role': 'teacher'} ,status=status.HTTP_200_OK)
        
            
    except :
        return json.dumps({"error": "User not found"})
       
        

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect



def signout(request):
    logout(request)
    messages.success(request,"you have successfully logged out")
    return redirect('home')


def generate_random_strings( length=10):
    characters = string.ascii_letters + string.digits + string.punctuation
    random_strings = []

    random_string = ''.join(random.choice(characters) for _ in range(length))
    #random_strings.append(random_string)

    return random_string
def register_users(request):
    # #now over here we will learn to make automatic sign ups 
# # automatic signups keh sath initial passwords and usernames jo hain woh must be stored somewhere
# # after this we will try to link SQL keh tables with django
# Open the Excel file
    excel_file = r'C:\\Users\\DELL\Desktop\\log\\login\\somehtingg.xlsx'
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    

    # Add the array elements to the document
    
        
    # Iterate through rows and create user accounts
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email = row[0]  # Assuming email is in the first column
        name=row[1]
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email Already Registered!!")
        else:
            username = email # Use a portion of the email as username
            password = generate_random_strings()  # Set a default password or generate one
            s=username+password
        # Create a new user
           

    # Save the document
            doc=Document()
                 
            user = User.objects.create_user(username=username, email=email, password=password, first_name=name)
            doc.add_paragraph(s)
            doc.save('usernames.docx')
            #user = User.objects.get(username=username)
            group, created = Group.objects.get_or_create(name='Instructor')
            user.groups.add(group)
    # Close the Excel file
    workbook.close()
    messages.success(request,"you have successfully registered all teachers")
    return redirect('home')


from django.db.models import Count
from django.db.models import Q

def save_role(request):
    if request.method == 'POST':
        
        usernames = request.POST.get('username')
        selected_role = request.POST.get('role')
        # Check if you are getting the values correctly
   
        group, created = Group.objects.get_or_create(name=selected_role)
        #user.groups.add(instructor_group)
        user=User.objects.filter(username=usernames)
        user.groups.add(selected_role)
   
        
    return HttpResponse('shyad')
main_selected_instructor=""

@api_view(['GET'])
def get_instructors(request):
    # set_g_instructor
    teachers = Teacher.objects.get(first_name=set_g_instructor)
    teacherss = Teacher.objects.all()
    user_groups = teachers.groups.all()
    if any(group.name == 'HOS' for group in user_groups):
        teacher_names = [teacher.first_name for teacher in teacherss]
        data = {
            'teachers_name': teacher_names,
            # 'qs': qs,
        }
        return JsonResponse(data, safe=False)
    else:
            data = {
                'teachers_name': teachers.name,
                # 'qs': qs,
            }
            return JsonResponse(data, safe=False)
    
@api_view(['GET','POST'])
def get_semester_year_combinations(request):
    teacher_course_id=request.data.get('teacher_course_id')
    sections = Section.objects.filter(teacher_course_id=teacher_course_id)
    semester_year_combinations = []

    for section in sections:
        semester_year_combinations.append({
            'semester': section.semester,
            'year': section.year,
            'section_name': section.section_name,
            'total_registered_students': section.registered_students,
            'total_feedbacks_received': section.feedbacks_received,
            'total_positive_comments': section.number_of_positive_comments,
            'total_constructive_comments': section.number_of_constructive_comments,
        })

    return JsonResponse(semester_year_combinations, safe=False)
@api_view(['GET','POST'])
def get_student_courses(request):
    teachersname=request.data.get('teachersname')
    sections = Section.objects.filter(teacher_course_id=teacher_course_id)
    semester_year_combinations = []

    for section in sections:
        semester_year_combinations.append({
            'semester': section.semester,
            'year': section.year,
            'section_name': section.section_name,
            'total_registered_students': section.registered_students,
            'total_feedbacks_received': section.feedbacks_received,
            'total_positive_comments': section.number_of_positive_comments,
            'total_constructive_comments': section.number_of_constructive_comments,
        })

    return JsonResponse(semester_year_combinations, safe=False)
from django.http import JsonResponse
from django.db.models import Count, Q


@api_view(['GET'])
def assign_role(request):
    # Retrieve groups with no members
    groups_with_no_members = Group.objects.annotate(num_members=Count('user')).filter(num_members=0)

    # Convert the result to a list of dictionaries containing only ID and name
    qs = list(groups_with_no_members.values('id', 'name'))

    exclude_groups = ["HOS", "HOD CS", "HOD SE", "HOD EE", "HOD humanities", "HOD FSM", "Admin"]

    # Create a Q object to filter users not in the exclude_groups
    exclude_condition = Q(groups__name__in=exclude_groups)

    # Get a list of users not in the exclude_groups
    users_not_in_exclude_groups = Teacher.objects.exclude(exclude_condition)

    # Extracting relevant fields from users_not_in_exclude_groups queryset
    # users_data = list(users_not_in_exclude_groups.values('id', 'name'))

    # Constructing response data
    data = {
        # 'users_not_in_exclude_groups': users_data,
        'qs': qs,
    }
    roles = [item['name'] for item in qs]
    data = {
        'roles': roles,
    }

    return JsonResponse(data, safe=False)

@api_view(['GET'])
def assign_role_person(request):
    qs=Group.objects.all()
    #selected_role = request.POST.get('role')
    #return render(request, 'authentication\\assigning_permission.html', {'qs': qs})
    
    groups_with_no_members = Group.objects.annotate(num_members=Count('user')).filter(num_members=0)

# Convert the result to a list of dictionaries
    qs = list(groups_with_no_members.values())
    
    exclude_groups = ["HOS", "HOD CS", "HOD SE", "HOD EE", "HOD humanities", "HOD FSM", "Admin"]

# Create a Q object to filter users not in the exclude_groups
    exclude_condition = Q(groups__name__in=exclude_groups)

# Get a list of users not in the exclude_groups
    users_not_in_exclude_groups = Teacher.objects.exclude(exclude_condition)
    users_data = list(users_not_in_exclude_groups.values('teacher_id', 'username'))

    data = {
        'users_not_in_exclude_groups': users_data,
        
    }
    return JsonResponse(data, safe=False)


@api_view(['POST'])
def save_data(request):
    if request.method == 'POST':
        data = request.data
        role = data.get('role')
        user = data.get('user')
        group, created = Group.objects.get_or_create(name=role)

           
        user = Teacher.objects.get(teacher_id=user)  # Assuming usernames are unique
        user.groups.add(group)
        return JsonResponse({'message': 'Data saved successfully'}, status=200)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
@api_view(['POST'])
def save_role(request):

    if request.method == 'POST':
        #mail = request.data.get('username')
        usernames = request.data.get('username')
        selected_role = request.data.get('role')
        
       #  group, created = Group.objects.get_or_create(name='Admin')
        #user.groups.add(instructor_group)
        #user.groups.add(group)
        
        group, created = Group.objects.get_or_create(name=selected_role)

           
        user = User.objects.get(username=usernames)  # Assuming usernames are unique
        user.groups.add(group)
        return Response({'message': 'User deleted'}, status=status.HTTP_200_OK)


set_g_instructor = ""  # Initialize the global variable outside the function
# global set_g_instructor

@api_view(['POST'])
def set_instructor(request):
    global set_g_instructor  # Declare that you are using the global variable
    set_g_instructor = request.data.get('newSelectedInstructor')

   

    return Response({'message': 'Value received successfully'})

 
def register_single_user(request):
    # #now over here we will learn to make automatic sign ups 
# # automatic signups keh sath initial passwords and usernames jo hain woh must be stored somewhere
# # after this we will try to link SQL keh tables with django
# Open the Excel file
    excel_file = r'C:\\Users\\DELL\Desktop\\log in system\\login\\somehting.xlsx'
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Iterate through rows and create user accounts
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email = row[0]  # Assuming email is in the first column
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email Already Registered!!")
        else:
            username = email # Use a portion of the email as username
            password = generate_random_strings()  # Set a default password or generate one

        # Create a new user
            User.objects.create_user(username=username, email=email, password=password)

    # Close the Excel file
    workbook.close()
    messages.success(request,"you have successfully registered all teachers")
    return redirect('home')
def change_pass(request):
    user = request.user
    if request.method=="POST":
        old_pass=request.POST['current_password']
        new_pass=request.POST['new_pass']
        new_pass_conf=request.POST['new_pass_conf']
        if user.check_password(old_pass) and new_pass == new_pass_conf:
            user.set_password(new_pass_conf)
            user.save()
        update_session_auth_hash(request, user)  # Update session to avoid automatic logout
        messages.success(request, 'Password changed successfully.')
        return redirect('home')  # Redirect to a profile page or wherever you prefer
    else:
        messages.error(request, 'Password change failed. Please check your inputs.')
        
    return render(request, 'authentication\changepassword.html')
from django.http import JsonResponse

# @api_view(['GET','POST'])
# def get_instructor_courses(request):
    
  
    
#     # Get the teacher based on the first name
#     teacher = get_object_or_404(Teacher, first_name=set_g_instructor)

#     # Get all courses associated with the teacher
#     teacher_courses = TeacherCourse.objects.filter(teacher=teacher)

#     # Extract course information with only course codes
#     courses_data = [{'course_code': tc.course.course_name} for tc in teacher_courses]

#     # You can customize the response data structure based on your requirements
#     response_data = {'courses': courses_data}

#     return JsonResponse(response_data)
@api_view(['GET', 'POST'])
def get_instructor_courses(request):
    global global_to
    global global_from
    global_from=request.data.get('selectedFrom')           
    global_to=request.data.get('selectedTo')
    # Filter sections by year and semester
    if request.data.get('selectedFrom')=='' and request.data.get('selectedTo')=='':
        sections = Section.objects.filter(year=g_year, semester=g_semester)
        #teacher_courses = TeacherCourse.objects.filter(teacher=teacher)


        # Filter sections further by subject code and teacher
        filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)

        # Get unique courses associated with filtered sections
        courses = Course.objects.filter(teachercourse__section__in=filtered_sections).distinct()

        # Extract course information with only course codes
        courses_data = [{'course_code': course.course_name} for course in courses]

        # You can customize the response data structure based on your requirements
        response_data = {'courses': courses_data}

        return Response(response_data)
    elif request.data.get('selectedTo')==request.data.get('selectedFrom'):
        To= request.data.get('selectedTo')
        start_semester, start_year = To.split()
        sections = Section.objects.filter(year=start_year, semester=start_semester)
        #teacher_courses = TeacherCourse.objects.filter(teacher=teacher)


        # Filter sections further by subject code and teacher
        filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)

        # Get unique courses associated with filtered sections
        courses = Course.objects.filter(teachercourse__section__in=filtered_sections).distinct()

        # Extract course information with only course codes
        courses_data = [{'course_code': course.course_name} for course in courses]

        # You can customize the response data structure based on your requirements
        response_data = {'courses': courses_data}

        return Response(response_data)
    else:
        From= request.data.get('selectedFrom')
        To= request.data.get('selectedTo')
        start_semester, start_year = From.split()
        end_semester, end_year = To.split()
        end_year=int(end_year)
        start_year=int(start_year)
        SEMESTER_ORDER = {
        'spring': 1,
        'summer': 2,
        'fall': 3,
        }

        years_range = list(range(start_year, end_year + 1))


        # Query sections
        sections = Section.objects.filter(
            Q(year__in=years_range) |
            Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
            Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
            Q(year__gt=start_year, year__lt=end_year)
        )
        # Get unique courses associated with filtered sections
        filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
        courses = Course.objects.filter(teachercourse__section__in=filtered_sections).distinct()

        # Extract course information with only course codes
        courses_data = [{'course_code': course.course_name} for course in courses]

        # You can customize the response data structure based on your requirements
        response_data = {'courses': courses_data}

        return Response(response_data)       




def generate_random_alphanumeric_string(length):
    characters = string.ascii_letters + string.digits
    random_string = ''.join(random.choice(characters) for _ in range(length))
    return random_string
# @api_view(['POST'])
# def create_user(request):
#     if request.method=="POST":
         
#         email = request.data.get('email')
#         name = request.data.get('name')
#         if User.objects.filter(username=email):
#             messages.error(request, "Username already exist! Please try some other username.")
#             return redirect('home')
#         password=generate_random_alphanumeric_string(10)
#         myuser=User.objects.create_user(email, email,password)
#         myuser.first_name=name
#         myuser.save()
#         messages.success(request,"the user has been successfuly registered")
#         if myuser is not None:
#             return Response({'message': 'you have created user','password':password}, status=status.HTTP_200_OK)
#         else:
#             return Response({'message': 'you have not created user'}, status=status.HTTP_200_OK)

            

from django.contrib.auth.models import User
@api_view(['POST'])
def delete_user(request):
    if request.method=="POST":
        email = request.data.get('username')
        user_to_delete = User.objects.get(username=email)
        # Delete the user
        user_to_delete.delete()
        return Response({'message': 'User deleted'}, status=status.HTTP_200_OK)


def assigning(request):
    teacher = Teacher.objects.get(first_name='Kashif Munir')
    hos_group, created = Group.objects.get_or_create(name="HOS")
    teacher.groups.add(hos_group)
# Add the teacher to the group
    # hos_group.members.add(teacher)

    return render(request, 'authentication\signup.html')

    
from django.db.models import Sum
#     return render(request, 'authentication\signup.html')
@api_view(['POST','GET'])
def years_of_teacher(request):
    try:
        # Assuming set_g_instructor is defined somewhere before this code

        # Get the Teacher object
        teacher = Teacher.objects.get(first_name=set_g_instructor)

        # Get the TeacherCourse objects for the teacher
        teacher_courses = TeacherCourse.objects.filter(teacher=teacher)

        # Initialize lists to store data
        semester_years = []
        positive_ones = []
        negative_ones = []

        # Iterate over each TeacherCourse
        for teacher_course in teacher_courses:
            # Get sections for this TeacherCourse
            sections = Section.objects.filter(teacher_course=teacher_course)

            # Sum of positive comments for each semester year
            positive_comments_by_year = sections.values('semester', 'year').annotate(
                total_positive_comments=Sum('number_of_positive_comments')
            )

            # Sum of negative comments for each semester year
            negative_comments_by_year = sections.values('semester', 'year').annotate(
                total_negative_comments=Sum('number_of_constructive_comments')
            )

            # Populate arrays with values
            for data in positive_comments_by_year:
                semester_years.append(f"{data['semester']} {data['year']}")
                positive_ones.append(data['total_positive_comments'])

            for data in negative_comments_by_year:
                negative_ones.append(data['total_negative_comments'])
        data = {
            'positive': positive_ones,
            'negative': negative_ones,
            'semester_year': semester_years,
        }

        return JsonResponse(data)

    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'})

    except TeacherCourse.DoesNotExist:
        return JsonResponse({'error': 'TeacherCourse not found'})
@api_view(['POST','GET'])
def feedback_saving(request):
    if request.method == 'POST':
        comment_text = request.data.get('comment')
        section_id = request.data.get('section_id')

        # Assuming Section model is imported and Section object is available
        try:
            section = Section.objects.get(pk=section_id)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=status.HTTP_404_NOT_FOUND)

        # Create and save the comment
        comment = Comment.objects.create(section=section, comments=comment_text)
        
        # Optionally, you can also save the category if available in the request data
        category = request.data.get('category')
        if category:
            comment.category = category
            comment.save()

        return Response({'message': 'Comment saved successfully'}, status=status.HTTP_201_CREATED)
    else:
        return Response({'error': 'Only POST method allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    

from django.db.models import Max


@api_view(['GET', 'POST'])
def get_instructor_coursen(request):
    # Get the Teacher object
    if request.data.get('selectedFrom')=='' and request.data.get('selectedTo')=='':
        set_g_instructor = request.data.get('selectedInstructor')
        global_course = request.data.get('selectedCourses')
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        course = Course.objects.get(course_name=global_course)
        teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)
        filtered_sections = Section.objects.filter(teacher_course=teacher_course, year=g_year, semester=g_semester)
        section_names = [section.section_name for section in filtered_sections]
        response_data = {'sectionn': section_names}
        return JsonResponse(response_data)
    elif request.data.get('selectedFrom')==request.data.get('selectedTo'):
        From=request.data.get('selectedFrom')
        gsem,gyear=From.split()
        teacher = request.data.get('selectedInstructor')
        course = request.data.get('selectedCourses')
        teacher = Teacher.objects.get(first_name=teacher)
        course = Course.objects.get(course_name=course)
        teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)
        sections = Section.objects.filter(teacher_course=teacher_course, year=gyear, semester=gsem)
        section_names = [section.section_name for section in sections]
        response_data = {'sectionn': section_names}
        return JsonResponse(response_data)
    else:
        From= request.data.get('selectedFrom')
        To= request.data.get('selectedTo')
        set_g_instructor = request.data.get('selectedInstructor')
        global_course = request.data.get('selectedCourses')
        start_semester, start_year = From.split()
        end_semester, end_year = To.split()
        end_year=int(end_year)
        start_year=int(start_year)
        SEMESTER_ORDER = {
        'spring': 1,
        'summer': 2,
        'fall': 3,
        }

        years_range = list(range(start_year, end_year + 1))


        sections = Section.objects.filter(
            Q(year__in=years_range) |
            Q(year=start_year, semester__in=SEMESTER_ORDER.keys(), semester__gte=start_semester) |
            Q(year=end_year, semester__in=SEMESTER_ORDER.keys(), semester__lte=end_semester) |
            Q(year__gt=start_year, year__lt=end_year)
        )
        # Get unique courses associated with filtered sections
        filtered_sections = sections.filter(teacher_course__teacher__first_name=set_g_instructor)
        filtered_sections = filtered_sections.filter(teacher_course__course__course_name=global_course)
        
        # From=request.data.get('selectedFrom')
        # To=request.data.get('selectedTo')
        # start_semester,start_year=From.split()
        # end_semester,end_year=To.split()
        # teacher = Teacher.objects.get(first_name=set_g_instructor)
        # course = Course.objects.get(course_name=global_course)
        # teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)
        # sections = Section.objects.filter(teacher_course=teacher_course, year=g_year, semester=g_semester)
        section_names = [section.section_name for section in filtered_sections]
        response_data = {'sectionn': section_names}
        return JsonResponse(response_data)
@api_view(['GET','POST'])
def return_section(request):
    try:
   
        # Create a JSON response
        response_data = {
            'section': global_section,
           
        }
    
     
        return JsonResponse(response_data, safe=False)
    except Exception as e:
        # Handle exceptions appropriately
        response_data = {
            'status': 'error',
            'message': f'Error: {str(e)}',
        }
        return JsonResponse(response_data, status=500, safe=False)
    

@api_view(['POST','GET'])
def category_wise(request):
    try:
        set_g_instructor
        resp = request.data.get('response')
        category_wise = global_cat
        section = global_section  # global section jo select huaa huaa hai
        # idhar se find the section
        # sabb find krke insert
        teacher = Teacher.objects.get(first_name=set_g_instructor)

        # Get the Course object
        course = Course.objects.get(course_name=global_course)

        # Get the TeacherCourse object
        teacher_course = TeacherCourse.objects.get(teacher=teacher, course=course)

        # Get the Section object
        section_instance = Section.objects.get(teacher_course=teacher_course, semester=g_semester, year=g_year, section_name=section)

        # Initialize a variable to store the total registered students
        # are we setting global section
        # kisi trhaa global section ki id i need
        comment = Respon.objects.create(section=section_instance, respon=resp, category=category_wise)
        comment.save()

        return JsonResponse({'success': 'data is stored'})
    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'})

    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'})

    except TeacherCourse.DoesNotExist:
        return JsonResponse({'error': 'TeacherCourse not found'})

    except Section.DoesNotExist:
        return JsonResponse({'error': 'Section not found'})




#things we need to do, jabb global section select ho toh waapis ID aajaye
#jabb id waapis aaajaye toh global set hojaye this is what we need to fix in jugaar time
#saray components ko ek traf le aaye so we can update
#
    



from itertools import cycle


@api_view(['POST','GET'])
def setFromDropdown(request):
    if request.method == 'POST':
        # Assuming set_g_instructor and global_course are sent in the POST request
        set_g_instructor = request.data.get('selectedInstructor')
       

        if set_g_instructor is None :
            return JsonResponse({'error': 'set_g_instructor are required'}, status=400)

        try:
            # Get the Teacher object
            teacher = Teacher.objects.get(first_name=set_g_instructor)
        except Teacher.DoesNotExist:
            return JsonResponse({'error': 'Teacher does not exist'}, status=404)

        # Get all courses taught by the teacher
        teacher_courses = TeacherCourse.objects.filter(teacher=teacher)

        # Initialize empty lists to store courses and sections
        courses = []
        sections = []

        # Iterate over teacher courses
        for teacher_course in teacher_courses:
            # Append course to the courses list
            courses.append({
                'course_name': teacher_course.course.course_name,
                'course_code': teacher_course.course.course_code
            })

            # Get all sections taught by the teacher for this course
            sections_queryset = Section.objects.filter(teacher_course=teacher_course)

            # Convert sections queryset to list of dictionaries
            sections.extend(list(sections_queryset.values()))
            
        if sections:
            semester_precedence = {'Spring': 0, 'Summer': 1, 'Fall': 2}

    # Sort sections by year and semester, considering semester precedence
            sections.sort(key=lambda x: (x['year'], semester_precedence[x['semester']]))

            min_year_section =sections[0] #min(sections, key=lambda x: (int(x['year']), x['semester']))
            max_year_section =  sections[-1]#max(sections, key=lambda x: (int(x['year']), x['semester']))
            # print("maximu year",max_year_section)
            # print("minimum year sections",min_year_section)
            # Generate the sequence of semester years
            semester_cycle = cycle(['Spring', 'Fall'])
            current_year = int(min_year_section['year'])
            current_semester = min_year_section['semester']
            sequence = []

            while True:
                sequence.append(f'{current_semester} {current_year}')

                if current_year == int(max_year_section['year']) and current_semester == max_year_section['semester']:
                    break

                current_semester = 'Spring' if current_semester == 'Fall' else 'Fall'

                if current_semester == 'Spring':
                    current_year += 1

            response_data = {
                'semesterYear': sequence
            }

            # print(sequence)

            return JsonResponse(response_data)

        else:
            response_data = {
                'semesterYear': []
            }

            # print(sequence)

            return JsonResponse(response_data)

    return JsonResponse({'error': 'Only POST method is supported'}, status=405)


@api_view(['POST','GET'])
def setToDropdown(request):
        if request.method == 'POST':
        # Assuming set_g_instructor and global_course are sent in the POST request
            set_g_instructor = request.data.get('selectedInstructor')
            # print('\n\n\nselectedInstructor: ',set_g_instructor)
            # print('\n\n\n')
            # global_course = request.data.get('global_course', None)

            if set_g_instructor is None :
                return JsonResponse({'error': 'set_g_instructor are required'}, status=400)

            try:
                # Get the Teacher object
                teacher = Teacher.objects.get(first_name=set_g_instructor)
            except Teacher.DoesNotExist:
                return JsonResponse({'error': 'Teacher does not exist'}, status=404)

            # Get all courses taught by the teacher
            teacher_courses = TeacherCourse.objects.filter(teacher=teacher)

            # Initialize empty lists to store courses and sections
            courses = []
            sections = []

            # Iterate over teacher courses
            for teacher_course in teacher_courses:
                # Append course to the courses list
                courses.append({
                    'course_name': teacher_course.course.course_name,
                    'course_code': teacher_course.course.course_code
                })

                # Get all sections taught by the teacher for this course
                sections_queryset = Section.objects.filter(teacher_course=teacher_course)

                # Convert sections queryset to list of dictionaries
                sections.extend(list(sections_queryset.values()))
                # for serction in sections:
                #     print(serction,"\n\n")
            # Find the minimum and maximum year and semester among the sections
            if sections:
                semester_precedence = {'Spring': 0, 'Summer': 1, 'Fall': 2}

        # Sort sections by year and semester, considering semester precedence
                sections.sort(key=lambda x: (x['year'], semester_precedence[x['semester']]))
                From= request.data.get('selectedFrom')
                start_semester, start_year = From.split()
                start_year = int(start_year)
                
                #min(sections, key=lambda x: (int(x['year']), x['semester']))
                max_year_section =  sections[-1]
                semester_cycle = cycle(['Spring', 'Fall'])
                current_year = start_year
                current_semester = start_semester
                sequence = []

                while True:
                    sequence.append(f'{current_semester} {current_year}')

                    if current_year == int(max_year_section['year']) and current_semester == max_year_section['semester']:
                        break

                    current_semester = 'Spring' if current_semester == 'Fall' else 'Fall'

                    if current_semester == 'Spring':
                        current_year += 1

                response_data = {
                    'semesterYear': sequence
                }

                #   

                return JsonResponse(response_data)

            else:
                response_data = {
                    'semesterYear': []
                }

                # print(sequence)

                return JsonResponse(response_data)

        return JsonResponse({'error': 'Only POST method is supported'}, status=405)
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
def send_email(first_name,receiver_email,password):
    # Setup the MIME
    # receiver_email=request.data.get('receiver_email')
    # receiver_email=request.data.get('username')
    # receiver_email=request.data.get('password')
    mail_content = f"Hello {first_name},These are your account credentials:\n\n Username:{receiver_email}\n\n Password:{password}\n\nYour account has been successfully created.\n\nThank you!\n"
    print("1done")
    subject='Account creation done'
    sender_email = 'i200908@nu.edu.pk'
    sender_pass = 'Blackshield@3128'
    print("1done")
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    # message['Cc'] = cc_email
    message['Subject'] = subject
    print("1done")
    message.attach(MIMEText(mail_content, 'plain'))
    # Create SMTP session for sending the mail
    session = smtplib.SMTP('smtp.gmail.com', 587)  # use gmail with port
    session.starttls()  # enable security
    session.login(sender_email, sender_pass)  # login with mail_id and password
    print("1done")
    text = message.as_string()
    session.sendmail(sender_email, [receiver_email], text)
    session.quit()
    print('Mail Sent')
    return True



@api_view(['POST'])
def create_user(request):
    if request.method == "POST":
        receiver_email = request.data.get('email')
        email = request.data.get('email')
        name = request.data.get('first_name')
        
        if Teacher.objects.filter(username=email).exists():
            messages.error(request, "Username already exists! Please try a different username.")
            return redirect('home')

        password = generate_random_alphanumeric_string(10)
        myuser = Teacher.objects.create_user(email=email, username=email, password=password)
        myuser.first_name = name  # Set the first name
        myuser.save()
        
        group, created = Group.objects.get_or_create(name='Instructor')
        myuser.groups.add(group)

        # Send email
        send_email(name, email, password)

        messages.success(request, "The user has been successfully registered")

        return Response({'message': 'You have created the user', 'password': password}, status=status.HTTP_200_OK)
    
@api_view(['POST'])
def check_user(request):
    try:
        set_g_instructor 
        print("srtginstruct:\n\n\n",set_g_instructor)
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
            else:
                teacher_course = TeacherCourse.objects.filter(teacher=teacher)

            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=g_semester, year=g_year)
                for sec in sections:
                    r=Response.objects.filter(section=sec)
                    if r:
                        something= True
                        response_data = {'value': something}
                    
            
            response_data = {'value':False}
            print("\n\n\nresponse_data: ",response_data)
            print("\n\n\n\n")
            return JsonResponse(response_data)
        elif global_to == global_from:
            sem,yea=global_to.split()
            teacher = Teacher.objects.get(first_name=set_g_instructor)
            teacher_course=""
            
            course = Course.objects.get(course_name=global_course)
            # Assuming g_semester and g_year are defined somewhere else
            teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
        
            feedbacks_received = 0
            
            for tech in teacher_course:
                sections = Section.objects.filter(teacher_course=tech, semester=sem, year=yea)
                for sec in sections:
                    r=Response.objects.filter(section=sec)
                    if r:
                        something= True
                        response_data = {'value': something}
                    
            response_data = {'value': False}
            return JsonResponse(response_data)

            
            
        else:
            # print("we in third condition")
          response_data = {'value': False}
          return JsonResponse(response_data)



    except Exception as e:
        # Handle other exceptions
        print(e)
        pass

    # Return a default response if conditions are not met or if an error occurs
    return JsonResponse({'error': 'Could not calculate feedbacks received'})
def check_same_logged(request):
    if logged_in==set_g_instructor:
        teacher = Teacher.objects.get(first_name=set_g_instructor)
        # = request.data.get('selectedInstructor')
        # Ensure global variables are defined properly
        global global_to
        global global_from

        if ('global_to' not in globals() or 'global_from' not in globals() or
            global_to == '' or global_to is None or global_from == '' or global_from is None):
            # Assuming set_g_instructor is defined somewhere else
            
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
           
            
            
            for tech in teacher_course:
                sections = Section.objects.get(section_name=global_section,teacher_course=tech, semester=g_semester, year=g_year)
                print("HERE ARE THE SECTINOS SELECTED")
                print(global_section)
                print(tech)
                print(sections)
                if sections:
                    try:
                        r=Respon.objects.get(section_id=sections.id)
                        print("\n\n sosmething ")
                        print(r)
                    except:
                        response_data = {'value': True}
                        print ("response_data: ",response_data)
                        return JsonResponse(response_data)

                    if r:
                        something= r.respon
                        response_data = {'value': something}
                        return JsonResponse(response_data)
                    else:
                        # something= r.respon
                        response_data = {'value': True}
                        print ("response_data: ",response_data)
                        return JsonResponse(response_data)
            response_data = {'value': False}
            return JsonResponse(response_data)
        elif (
            global_to ==  global_from):
            # Assuming set_g_instructor is defined somewhere else
            sem,yea=global_to.split()
            teacher_course=""
            if global_course!="":
                course = Course.objects.get(course_name=global_course)
                # Assuming g_semester and g_year are defined somewhere else
                teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
           
            
            
            for tech in teacher_course:
                sections = Section.objects.get(section_name=global_section,teacher_course=tech, semester=sem, year=yea)
                if sections:
                    r=Respon.objects.get(section=sections)
                    if r:
                        print(r)
                        something= r.respon
                        response_data = {'value': something}
                        return JsonResponse(response_data)
                    else:
                        # something= r.respon
                        response_data = {'value': True}
                        return JsonResponse(response_data)
            response_data = {'value': False}
            return JsonResponse(response_data)
                        
    else:
        response_data = {'value': False}
        return JsonResponse(response_data)
        # teacher = Teacher.objects.get(first_name=set_g_instructor)
        # # = request.data.get('selectedInstructor')
        # # Ensure global variables are defined properly
       

        # if ('global_to' not in globals() or 'global_from' not in globals() or
        #     global_to == '' or global_to is None or global_from == '' or global_from is None):
        #     # Assuming set_g_instructor is defined somewhere else
            
        #     teacher_course=""
        #     if global_course!="":
        #         course = Course.objects.get(course_name=global_course)
        #         # Assuming g_semester and g_year are defined somewhere else
        #         teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
           
            
            
        #     for tech in teacher_course:
        #         sections = Section.objects.get(section_name=global_section,teacher_course=tech, semester=g_semester, year=g_year)
        #         if sections:
        #             r=Respon.objects.get(section=sections)
        #             if r:
        #                 something= r.respon
        #                 response_data = {'value': something}
        #                 return JsonResponse(response_data)
                   
        #     response_data = {'value': False}
        #     return JsonResponse(response_data)
        # elif (
        #     global_to ==  global_from):
        #     # Assuming set_g_instructor is defined somewhere else
        #     sem,yea=global_to.split()
        #     teacher_course=""
        #     if global_course!="":
        #         course = Course.objects.get(course_name=global_course)
        #         # Assuming g_semester and g_year are defined somewhere else
        #         teacher_course = TeacherCourse.objects.filter(teacher=teacher, course=course)
           
            
            
        #     for tech in teacher_course:
        #         sections = Section.objects.get(section_name=global_section,teacher_course=tech, semester=sem, year=yea)
        #         if sections:
        #             r=Respon.objects.get(section=sections)
        #             if r:
        #                 something= r.respon
        #                 response_data = {'value': something}
        #                 return JsonResponse(response_data)
                    
        #     response_data = {'value': False}
        #     return JsonResponse(response_data)

                    
                    
                    
            
            
